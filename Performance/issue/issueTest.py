import requests
import json
import time
from datetime import datetime
import openpyxl
import re
from tqdm import tqdm
import os
'''
调用接口，去拿问题信息
'''
# 直接可以用开始时间、结束时间拿到所有要的信息
def issueno_get(StartDate,Enddate,excel_save_path):
    # 爬取数据
    Url = 'https://service.chanjet.com/partner_api/issue/search'
    params ={'starttime':f'{StartDate}',
           'endtime':f'{Enddate}'
           }
    Headers = {'Accept': '*/*',
               'Accept-Encoding': 'gzip, deflate, br',
               'Accept-Language': 'zh-CN,zh;q=0.9',
               'Referer': 'https://www.chanjetvip.com/partner/issue/search',
               'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36'
               }
    cookies = {'path': '/; domain=.chanjet.com; HttpOnly',
               '_identify': '647b7394740f4cfd6e1220ae559e2b2891734aa5c3508aada8b892c21b2b1f87a%3A2%3A%7Bi%3A0%3Bs%3A9%3A%22_identify%22%3Bi%3A1%3Bs%3A50%3A%22%5B%7B%22%24oid%22%3A%225e0d4a5d90fa1924b908803d%22%7D%2Cnull%2C2592000%5D%22%3B%7D',
               'expires': 'Thu, 21-Sep-2023 01:38:56 GMT',
               'Max-Age': '2592000',
               'domain': '.chanjetvip.com'
               }
    response = requests.get(url=Url, params=params, headers=Headers, cookies=cookies)
    dictres = json.loads(response.text)

    # excel写入第一行
    data_title = ['提问人','公司名称','问题号','产品分类','产品线','产品版本','产品模块','问题状态','处理工程师','问题类型','数据库类型','答复方式','数据方式','提问时间','首次回复时间','是否转研发','是否需要工程师掌握'] # 此处是要写入的数据
    wb = openpyxl.Workbook()
    ws = wb.create_sheet(title="Sundata",index=0)
    # 写数据函数cell,cell中column和row至少为1
    for i in range(1,len(data_title)+1):
        ws.cell(column=i, row=1, value=data_title[i-1])
    # 写入下面的数据
    if dictres["data"]:
        print(dictres["data"])
        if 'pagetotal' in dictres["data"].keys():
            pagemax = int(dictres["data"]['pagetotal'])
            counter = 1
            with tqdm(total=pagemax,iterable=1) as pbar:
                for i in range(1,pagemax+1):
                    time.sleep(1)
                    paramsnew={'starttime': f'{StartDate}',
                     'endtime': f'{Enddate}',
                     'page':f'{i}'
                     }
                    response = requests.get(url=Url, params=paramsnew, headers=Headers, cookies=cookies)
                    page_data =  json.loads(response.text)
                    # pageissuelist= [i["issueno"]for i in page_data["data"]["data"]]
                    # 构造需要的数据
                    data_frame =[[i["creator"]["name"],i["raise_org"]["name"],i["issueno"],i["product"],'软件包',i["pverion"],i["pmodule"],i["state"],i["supporter"]["name"],i["qtype"],i["dbtype"],i["replytype"],i["data_source"],i["created_at"],get_first_reply_time(i),get_first_turn_time(i),is_turn(i)[0],is_turn(i)[1]] for i in page_data["data"]["data"] ]
                    for k in data_frame:
                        for j in range(1, len(k)+1):
                            ws.cell(column=j,row=counter+1,value=k[j-1])
                        counter+=1
                    pbar.update(1)
                pbar.close()
                wb.save(excel_save_path)
        else:
            print('没查到数据')
    else:
        print('登陆信息有误')
'''
修正构造list里没有的信息
'''
# 获取首次回复的时间
def get_first_reply_time(alldata):
    flowlogs = alldata["flowlogs"]
    if len(flowlogs) == 1:
        return '无'
    else:
        i = 1
        while i < len(flowlogs) + 1:
            if flowlogs[len(flowlogs) - i]["type"] == 1:
                return create_time_formate(flowlogs[len(flowlogs) - i]["created_at"])
            else:
                i += 1
        return '无'
# 获取第一次转研发时间
def get_first_turn_time(alldata):
    if 'turn_dev_time' in alldata.keys():
        return timeformat(alldata["turn_dev_time"])
    else:
        return '无'
# 获取是否转研发
def is_turn(alldata):
    if 'advice_support' in alldata.keys():
        if alldata["advice_support"] == 1:
            return ['是','是']
        else:
            return ['是','否']
    else:
        return ['否','否']
# 获取创建时间（按格式来修复时间状态）
def create_time_formate(create_time):
    nowtime = int(time.time())
    if '前' not in create_time:
        return create_time
    else:
        if '秒' in create_time:
            time_degree = int(str(create_time).replace('秒前',''))
            return timeformat(nowtime-time_degree)
        elif '分钟' in create_time:
            time_degree = int(str(create_time).replace('分钟前', ''))
            return timeformat(nowtime - time_degree*60)
        elif '小时' in create_time:
            time_degree = int(str(create_time).replace('小时前', ''))
            return timeformat(nowtime - time_degree*3600)
        else:
            return create_time
'''
基础类通用函数
'''
# 时间格式函数
def timeformat(timeint):
    ti_shift1 = datetime.fromtimestamp(timeint)
    t1_return = ti_shift1.strftime("%Y-%m-%d %H:%M:%S")
    return t1_return
# 计算时间差的函数
def caculatetime(timeint):
    # ti_shift1 = datetime.fromtimestamp(timeint)
    timelist = convert_seconds(timeint)
    timeresult = str(timelist[0]) + '天' + str(timelist[1]) + '时' + str(timelist[2]) + '分' + str(timelist[3]) + '秒'
    return timeresult
def convert_seconds(seconds):
    days = seconds // 86400
    hours = (seconds - days * 86400) // 3600
    minutes = (seconds - days * 86400 - hours * 3600) // 60
    seconds = seconds - days * 86400 - hours * 3600 - minutes * 60
    return [days, hours, minutes, seconds]
'''
校验函数
'''
# 检验输入是否日期
def is_valid_datetime(string):
    pattern = "^\\d{4}-\\d{2}-\\d{2}"
    if re.match(pattern, string):
        return True
    else:
        return False
# 修正路径：
def repath(path):
    newpath = path.replace('/','\\')
    return newpath
# 校验是否是一个合法的地址信息
def is_exists_path(path):
    path = repath(path)
    folder =  os.path.exists(path)
    if not folder:
        return False
    else:
        return True
# 校验开始时间和结束时间哪个大
def compare_time(stime,etime):
    stime_int = int(datetime.strptime(stime, "%Y-%m-%d").timestamp())
    etime_int = int(datetime.strptime(etime, "%Y-%m-%d").timestamp())
    if stime_int > etime_int:
        return False
    else:
        return True

def Excel_export():
# 第一步
    try:
        Starttime = input("请输入查询的起始时间：")
        while not is_valid_datetime(Starttime):
            Starttime = input("请输入合理的时间格式：")
        Endtime = input("请输入查询的结束时间：")
        while not is_valid_datetime(Endtime):
            Endtime = input("请输入合理的时间格式：")
        while not compare_time(Starttime,Endtime):
            Endtime = input("开始时间不能晚于结束时间,请重新输入结束时间：")
        Savepath = input("请输入输出Excel数据的路径：（不需要输入输出的文件名）")
        while not is_exists_path(Savepath):
            Savepath =  input("请输入合法且存在的路径：")
        Saveexcel = Savepath.rstrip('/\\')+r'/output.xls'
        issueno_info = issueno_get(Starttime,Endtime,repath(Saveexcel))
        print(f'输出完成，Excel输出路径为:{Saveexcel}')
    except Exception as e:
        print('错误：',e)

if __name__ == '__main__':
    # Starttime = input("请输入查询的起始时间：")
    # Endtime = input("请输入查询的结束时间：")
    # Savepath = input("输出数据的路径：")
    # stime ='2023-8-22'
    # etime = '2023-8-22'
    # excel_save_path = r'I:\test.xls'
    # issueno_info = issueno_get(stime,etime,excel_save_path)
    Excel_export()

